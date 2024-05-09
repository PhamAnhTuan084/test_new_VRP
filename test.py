import warnings
import math
warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
from geopy import distance
import folium
import osmnx as ox
import networkx as nx
from shapely.geometry import Polygon
from shapely.geometry import LineString, Point
import random
from geopy.distance import geodesic
import streamlit as st
from streamlit_folium import folium_static
import base64
from io import BytesIO
from openpyxl.workbook import Workbook

def process_uploaded_files(uploaded_files):
    dataframes = {}
    data = None

    for idx, file in enumerate(uploaded_files):
        df = pd.read_excel(file)

        filename_without_extension = file.name.split('.')[0]
        dataframes[filename_without_extension] = df

        if idx == 0:
            data = df.copy()

    return dataframes, data

def create_square_map(group_1_df, min_lat, max_lat, min_lon, max_lon):
    # Tính toán tọa độ của tâm hình vuông
    center_lat = (min_lat + max_lat) / 2
    center_lon = (min_lon + max_lon) / 2

    # Tạo bản đồ mới với tọa độ tâm làm trung tâm và phóng to sao cho hình vuông nằm hoàn toàn trong bản đồ
    baby_map = folium.Map(location=[center_lat, center_lon], zoom_start=12)

    # Thêm marker cho mỗi điểm trong group_1_df
    for index, row in group_1_df.iterrows():
        popup_content = f"Order: {index+1}<br>OutletID: {row['OutletID']}<br>OutletName: {row['OutletName']}<br>Latitude: {row['Latitude']}<br>Longitude: {row['Longitude']}<br>tier: {row['tier']}"
        folium.Marker(location=[row['Latitude'], row['Longitude']], popup=folium.Popup(popup_content, max_width=300)).add_to(baby_map)

    # Chọn màu ngẫu nhiên từ danh sách các màu
    colors = ['black', 'beige', 'lightblue', 'gray', 'blue', 'darkred', 'lightgreen', 'purple', 'red', 'green', 'lightred', 'darkblue', 'darkpurple', 'cadetblue', 'orange', 'pink', 'lightgray', 'darkgreen', 'pink', 'yellow', 'purple']

    random_color = random.choice(colors)

    # Tạo hình vuông bao quanh các điểm với màu được chọn ngẫu nhiên
    folium.Rectangle(bounds=[(min_lat, min_lon), (max_lat, max_lon)], color=random_color, fill=True, fill_opacity=0.2).add_to(baby_map)

    return baby_map

def remove_outliers_iqr(data, factor=1.5):
    q1 = np.percentile(data, 25)
    q3 = np.percentile(data, 75)
    iqr = q3 - q1
    lower_bound = q1 - factor * iqr
    upper_bound = q3 + factor * iqr
    return data[(data >= lower_bound) & (data <= upper_bound)]

def find_nearest_corner_point(min_lat, min_lon, max_lat, max_lon, group_1_df):
    corners = [(min_lat, min_lon), (min_lat, max_lon), (max_lat, min_lon), (max_lat, max_lon)]

    nearest_points = {}

    for corner in corners:
        min_distance = np.inf
        nearest_point = None

        for index, row in group_1_df.iterrows():
            distance = np.sqrt((corner[0] - row['Latitude'])**2 + (corner[1] - row['Longitude'])**2)
            if distance < min_distance:
                min_distance = distance
                nearest_point = row

        nearest_points[corner] = {'point': nearest_point, 'distance': min_distance}

    min_distance_data = min(nearest_points.items(), key=lambda x: x[1]['distance'])

    return min_distance_data

def create_final_df(filtered_df, min_distance_data, no_outlet):
    # Khởi tạo final_df1 với cột dữ liệu tương tự như filtered_df
    final_df1 = pd.DataFrame(columns=filtered_df.columns)

    # Thêm điểm gần nhất vào final_df1
    final_df1 = pd.concat([final_df1, min_distance_data[1]['point'].to_frame().T], ignore_index=True)

    # Tính khoảng cách giữa các điểm trong filtered_df và điểm gần nhất
    filtered_df['distance_to_min'] = filtered_df.apply(lambda row: np.linalg.norm(np.array((row['Latitude'], row['Longitude'])) - np.array(min_distance_data[0])), axis=1)

    # Sắp xếp filtered_df theo khoảng cách tới điểm gần nhất
    filtered_df_sorted = filtered_df.sort_values(by='distance_to_min')

    # Tìm điểm kế tiếp ngắn nhất từ min_distance_data[1]['point']
    current_point = min_distance_data[1]['point']

    # Lặp để thêm điểm cho đến khi final_df1 có đủ 30 điểm
    while len(final_df1) < no_outlet:
        # Lấy điểm kế tiếp ngắn nhất
        next_point_index = filtered_df_sorted.index[0]
        next_point = filtered_df_sorted.iloc[0]

        # Loại bỏ điểm đã chọn khỏi filtered_df_sorted
        filtered_df_sorted = filtered_df_sorted.drop(next_point_index)

        # Nếu điểm kế tiếp không trùng với điểm hiện tại, thêm vào final_df1
        if next_point_index != current_point.name:
            final_df1 = pd.concat([final_df1, pd.DataFrame([next_point], columns=final_df1.columns)], ignore_index=True)

        # Cập nhật điểm hiện tại là điểm kế tiếp đã chọn
        current_point = next_point

    return final_df1

def draw_small_square(final_df1, map, random_color):
    # Tính toán tọa độ của hình vuông bao quanh các điểm trong final_df
    min_lat_final1 = final_df1['Latitude'].min()
    max_lat_final1 = final_df1['Latitude'].max()
    min_lon_final1 = final_df1['Longitude'].min()
    max_lon_final1 = final_df1['Longitude'].max()

    # Chọn màu ngẫu nhiên từ danh sách các màu
    # colors = ['black', 'beige', 'lightblue', 'gray', 'blue', 'darkred', 'lightgreen', 'purple', 'red', 'green', 'lightred', 'white', 'darkblue', 'darkpurple', 'cadetblue', 'orange', 'pink', 'lightgray', 'darkgreen', 'pink', 'yellow', 'purple']
    # random_color = random.choice(colors)

    # Tạo hình vuông bao quanh các điểm trong final_df với màu được chọn ngẫu nhiên
    folium.Rectangle(bounds=[(min_lat_final1, min_lon_final1), (max_lat_final1, max_lon_final1)], color=random_color, fill=True, fill_opacity=0.2).add_to(map)

    return map

def euclidean_distance(point1_coords, point2_coords):
    return math.sqrt((point2_coords['Longitude'] - point1_coords['Longitude'])**2 + (point2_coords['Latitude'] - point1_coords['Latitude'])**2)

def calculate_distance_euclidean_two_points(point1_coords, point2_coords, filtered_df):
    distances = {}
    for index, row in filtered_df.iterrows():
        distance = euclidean_distance(point1_coords, {'Longitude': row['Longitude'], 'Latitude': row['Latitude']})
        distance += euclidean_distance(point2_coords, {'Longitude': row['Longitude'], 'Latitude': row['Latitude']})
        distances[index] = distance
    return distances

def create_final_df2(filtered_sorted_df, closest_points, no_oulet):
    final_df2 = pd.DataFrame(columns=filtered_sorted_df.columns)
    final_df2 = pd.concat([final_df2, closest_points.iloc[0].to_frame().T], ignore_index=True)

    # Tính khoảng cách giữa các điểm trong filtered_sorted_df và điểm closest_points[0]
    filtered_sorted_df['distance_to_closest'] = filtered_sorted_df.apply(lambda row: np.linalg.norm(np.array((row['Latitude'], row['Longitude'])) - np.array((closest_points.iloc[0]['Latitude'], closest_points.iloc[0]['Longitude']))), axis=1)

    # Sắp xếp filtered_df theo khoảng cách tới điểm gần nhất
    filtered_distance_sorted = filtered_sorted_df.sort_values(by='distance_to_closest')

    current_point = closest_points.iloc[0]

    while len(final_df2) < no_oulet and current_point is not None:
        # Lấy điểm kế tiếp ngắn nhất
        next_point_index = filtered_distance_sorted.index[0]
        next_point = filtered_distance_sorted.iloc[0]

        # Loại bỏ điểm đã chọn khỏi filtered_distance_sorted
        filtered_distance_sorted = filtered_distance_sorted.drop(next_point_index)

        # Nếu điểm kế tiếp không trùng với điểm hiện tại, thêm vào final_df2
        if next_point_index != current_point.name:
            final_df2 = pd.concat([final_df2, pd.DataFrame([next_point], columns=final_df2.columns)], ignore_index=True)

        # Cập nhật điểm hiện tại là điểm kế tiếp đã chọn
        current_point = next_point

    return final_df2

def Create_square(cleaned_data, no_oulet, new_map, random_color):
    min_lat = cleaned_data['Latitude'].min()
    max_lat = cleaned_data['Latitude'].max()
    min_lon = cleaned_data['Longitude'].min()
    max_lon = cleaned_data['Longitude'].max()

    # new_map = create_square_map(cleaned_data, min_lat, max_lat, min_lon, max_lon)

    min_distance_data = find_nearest_corner_point(min_lat, min_lon, max_lat, max_lon, cleaned_data)

    filtered_df = cleaned_data.drop(min_distance_data[1]['point'].name)
    final_df1 = create_final_df(filtered_df, min_distance_data, no_oulet)

    new_map = draw_small_square(final_df1, new_map, random_color)

    i = 1

    final_df1['SRD'] = i

    all_data = pd.DataFrame(columns=final_df1.columns)
    # all_data = all_data.append(final_df1, ignore_index=True)
    all_data = pd.concat([all_data, final_df1], ignore_index=True)

    filtered_sorted_df = filtered_df[~filtered_df['OutletID'].isin(final_df1['OutletID'])]

    i = i + 1

    while not filtered_sorted_df.empty:
        # print(i)
        last_two_rows = final_df1.tail(2)
        # Kiểm tra kích thước của filtered_sorted_df
        if len(filtered_sorted_df) <= 30:
            filtered_sorted_df['SRD'] = i
            # all_data = all_data.append(filtered_sorted_df, ignore_index=True)
            all_data = pd.concat([all_data, filtered_sorted_df], ignore_index=True)
            break

        distances_between_two_points = calculate_distance_euclidean_two_points(last_two_rows.iloc[0], last_two_rows.iloc[1], filtered_sorted_df)
        distances_series = pd.Series(distances_between_two_points)

        closest_points_indices = distances_series.nsmallest(1).index
        closest_points = filtered_sorted_df.loc[closest_points_indices]

        filtered_sorted_df = filtered_sorted_df[~filtered_sorted_df['OutletID'].isin(closest_points['OutletID'])]

        final_df1 = create_final_df2(filtered_sorted_df, closest_points, no_oulet)
        final_df1['SRD'] = i
        # all_data = all_data.append(final_df1, ignore_index=True)
        all_data = pd.concat([all_data, final_df1], ignore_index=True)
        new_map = draw_small_square(final_df1, new_map, random_color)

        i = i + 1

        filtered_sorted_df = filtered_sorted_df[~filtered_sorted_df['OutletID'].isin(final_df1['OutletID'])]

    return all_data, new_map

def Create_RD(all_data):
  # Khởi tạo một biến để lưu trữ số của mỗi loại RD
  rd_counts = {}

  # Duyệt qua từng hàng trong DataFrame
  for index, row in all_data.iterrows():
      rd = row['SRD']
      if rd not in rd_counts:
          rd_counts[rd] = 1
      else:
          rd_counts[rd] += 1
      # Tạo giá trị cho cột mới dựa trên RD và số lượng đã đếm
      all_data.at[index, 'RD'] = f"{rd}.{rd_counts[rd]}"

  return all_data


def calculate_distance(point_coords, filtered_df):
    # Convert point_coords to a Point object
    point = Point(point_coords[::-1])  # Reverse the order of coordinates

    # Calculate distance using Shapely and store in a new column
    filtered_df['distance_to_point'] = filtered_df.apply(
        lambda row: point.distance(Point(row['Longitude'], row['Latitude'])),
        axis=1
    )

    # Find the closest point
    closest_point = filtered_df.loc[filtered_df['distance_to_point'].idxmin()]

    return closest_point['distance_to_point'], closest_point

def draw_optimal_path(visited_points, new_map, G, random_color, group_feature):
    # Initialize cal_distance and cal_time with default values
    cal_distance = 0
    cal_time = 0
    
    # Extract the last two points from visited_points
    visited_points_df = pd.DataFrame(visited_points.tail(2))
    last_point = (visited_points_df.iloc[-2]['Latitude'], visited_points_df.iloc[-2]['Longitude'])
    final_point = (visited_points_df.iloc[-1]['Latitude'], visited_points_df.iloc[-1]['Longitude'])

    # Find the optimal path using OSMnx
    start_node = ox.distance.nearest_nodes(G, last_point[1], last_point[0])
    destination_node = ox.distance.nearest_nodes(G, final_point[1], final_point[0])

    optimal_path_nodes = ox.shortest_path(G, start_node, destination_node, weight='length')

    if optimal_path_nodes is not None:
        optimal_path_coordinates = [(G.nodes[node]['y'], G.nodes[node]['x']) for node in optimal_path_nodes]

        # Calculate distance and time
        cal_distance = ox.distance.great_circle_vec(last_point[0], last_point[1], final_point[0], final_point[1])
        cal_time = cal_distance / (20 / 3.6)  # Assuming speed is in m/s

        # Vẽ đường dẫn tối ưu giữa hai điểm cuối cùng
        poly_line = folium.PolyLine(optimal_path_coordinates, color=random_color, weight=2.5, opacity=1)
        poly_line.add_to(group_feature)

        # # Thêm điểm popup cho điểm thứ hai từ cuối
        # last_point_popup = f"Order: {visited_points_df.index[-2] + 1}<br>OutletID: {visited_points_df['OutletID'].iloc[-2]}<br>OutletName: {visited_points_df['OutletName'].iloc[-2]}<br>Latitude: {last_point[0]}<br>Longitude: {last_point[1]}<br>SRD: {visited_points_df['SRD'].iloc[-2]}<br>RD: {visited_points_df['RD'].iloc[-2]}<br>Flow: {visited_points_df['Flow'].iloc[-2]}<br>tier: {visited_points_df['tier'].iloc[-2]}"
        # last_marker = folium.Marker(location=[last_point[0], last_point[1]], popup=folium.Popup(last_point_popup, max_width=300), icon=folium.Icon(color=random_color))
        # last_marker.add_to(group_feature)

        # # Thêm điểm popup cho điểm cuối cùng
        # final_point_popup = f"Order: {visited_points_df.index[-1] + 1}<br>OutletID: {visited_points_df['OutletID'].iloc[-1]}<br>OutletName: {visited_points_df['OutletName'].iloc[-1]}<br>Latitude: {final_point[0]}<br>Longitude: {final_point[1]}<br>SRD: {visited_points_df['SRD'].iloc[-1]}<br>RD: {visited_points_df['RD'].iloc[-1]}<br>Flow: {visited_points_df['Flow'].iloc[-1]}<br>tier: {visited_points_df['tier'].iloc[-1]}"
        # final_marker = folium.Marker(location=[final_point[0], final_point[1]], popup=folium.Popup(final_point_popup, max_width=300), icon=folium.Icon(color=random_color))
        # final_marker.add_to(group_feature)

    else:
        print("No path found between", start_node, "and", destination_node)

    return cal_distance, cal_time, new_map

def distance(lat1, lon1, lat2, lon2):
    return geodesic((lat1, lon1), (lat2, lon2)).meters

def calculate_distance_between_two_points(point1_coords, point2_coords, graph, filtered_df):
    # Get coordinates of point 1 and point 2
    point1_coords = (point1_coords['Latitude'], point1_coords['Longitude'])
    point2_coords = (point2_coords['Latitude'], point2_coords['Longitude'])

    # Find the closest nodes to point 1 and point 2
    closest_node1 = ox.distance.nearest_nodes(graph, point1_coords[1], point1_coords[0])
    closest_node2 = ox.distance.nearest_nodes(graph, point2_coords[1], point2_coords[0])

    # Find the nearest nodes for all destinations
    destinations_nodes = {}
    for index, row in filtered_df.iterrows():
        dest_node = ox.distance.nearest_nodes(graph, row['Longitude'], row['Latitude'])
        destinations_nodes[index] = dest_node

    # Calculate shortest paths for all destinations
    shortest_paths = {}
    for index, dest_node in destinations_nodes.items():
        try:
            distance1 = nx.shortest_path_length(graph, closest_node1, dest_node, weight='length')
            distance2 = nx.shortest_path_length(graph, closest_node2, dest_node, weight='length')
            shortest_paths[index] = distance1 + distance2
        except nx.NetworkXNoPath:
            shortest_paths[index] = float('inf')  # Assign a large value for unreachable nodes

    return shortest_paths

def find_nearest_point(visited_points, closest_points):
    # Lấy ra hai dòng cuối cùng từ visited_points
    last_two_rows = visited_points.tail(2)

    # Tạo tam giác từ hai điểm cuối cùng trong last_two_rows
    points = pd.DataFrame({'Latitude': [last_two_rows.iloc[0]['Latitude'], last_two_rows.iloc[1]['Latitude']],
                           'Longitude': [last_two_rows.iloc[0]['Longitude'], last_two_rows.iloc[1]['Longitude']]})

    # Khởi tạo các biến để lưu thông tin của tam giác nhỏ nhất
    min_perimeter = float('inf')
    min_perimeter_outlet_info = None

    # Duyệt qua mỗi điểm trong closest_points
    for index, row in closest_points.iterrows():
        # Thêm điểm thứ ba vào tam giác
        point_df = pd.DataFrame({'Latitude': [row['Latitude']], 'Longitude': [row['Longitude']]})
        points = pd.concat([points, point_df], ignore_index=True)
        triangle = Polygon(points)

        # Tính chu vi của tam giác
        perimeter = triangle.length

        # So sánh với chu vi nhỏ nhất đã tìm thấy
        if perimeter < min_perimeter:
            min_perimeter = perimeter
            min_perimeter_outlet_info = {'OutletID': row['OutletID'], 'OutletName': row['OutletName'],
                                         'Latitude': row['Latitude'], 'Longitude': row['Longitude'],
                                         'SRD': row['SRD'], 'RD': row['RD'], 'Flow': row['Flow'], 'tier': row['tier']}
            
        # Xóa điểm thứ ba để chuẩn bị cho lần duyệt tiếp theo
        points = points[:-1]

    # Tạo DataFrame từ min_perimeter_outlet_info
    min_perimeter_outlet_df = pd.concat([pd.DataFrame(min_perimeter_outlet_info, index=[0])])

    return min_perimeter_outlet_df

def create_path(group_1_df, G, new_map, random_color, i):
    min_lat = group_1_df['Latitude'].min()
    max_lat = group_1_df['Latitude'].max()
    min_lon = group_1_df['Longitude'].min()
    max_lon = group_1_df['Longitude'].max()

    # Sử dụng hàm để tạo bản đồ
    baby_map = create_square_map(group_1_df, min_lat, max_lat, min_lon, max_lon)
    
    # Sử dụng để tìm điểm gần góc
    min_distance_data = find_nearest_corner_point(min_lat, min_lon, max_lat, max_lon, group_1_df)

    # Lọc data
    filtered_df = group_1_df.drop(min_distance_data[1]['point'].name)
    
    # Assuming min_distance_data is defined somewhere in your code
    nearest_point_coords = min_distance_data[0]

    # Call the function to calculate distances and find the closest point
    closest_distance, closest_point = calculate_distance(nearest_point_coords, filtered_df)

    # print("Closest distance:", closest_distance)
    # print("Closest point:", closest_point)   

    start_point = min_distance_data[1]['point']

    # Tạo DataFrame rỗng để lưu các điểm đã thăm
    visited_points = pd.DataFrame(columns=['OutletID', 'OutletName', 'Latitude', 'Longitude', 'SRD', 'RD', 'Flow', 'tier'])

    # Thêm start_point vào DataFrame
    visited_points = pd.concat([visited_points, start_point.to_frame().T], ignore_index=True)
    visited_points = pd.concat([visited_points, closest_point.to_frame().T], ignore_index=True)    
    
    total_distance = 0
    total_time = 0
    filtered_sorted_df = group_1_df[~group_1_df['OutletID'].isin(visited_points['OutletID'])]

    group_flow = group_1_df['flow'].iloc[0] 
    group_feature = folium.FeatureGroup(name=group_flow + '-' + str(i)).add_to(new_map)

    cal_distance, cal_time, new_map = draw_optimal_path(visited_points, new_map, G, random_color, group_feature)
    total_distance += cal_distance
    total_time += cal_time
        
    while not filtered_sorted_df.empty:
        last_row = visited_points.tail(1).iloc[0]
        # Tâm của hình tròn
        center_lat, center_lon = last_row['Latitude'], last_row['Longitude']

        # Bán kính ban đầu của hình tròn
        radius = 100

        # Lặp cho đến khi tìm được ít nhất một điểm hoặc không thể tăng bán kính nữa
        while True:
            # Lọc dữ liệu
            filtered_data = []
            for index, row in filtered_sorted_df.iterrows():
                point_lat, point_lon = row['Latitude'], row['Longitude']
                if distance(center_lat, center_lon, point_lat, point_lon) <= radius:
                    filtered_data.append(row)

            # Tạo DataFrame từ dữ liệu lọc
            filtered_df_within_circle = pd.DataFrame(filtered_data)

            # Kiểm tra nếu có ít nhất một điểm trong hình tròn
            if len(filtered_df_within_circle) > 0:
                break
            
            # Nếu không có điểm nào và bán kính đã tăng lên, tăng bán kính thêm 100m và tiếp tục lặp
            radius += 100

        # In ra filtered_df_within_circle
        print(filtered_df_within_circle)

        last_two_rows = visited_points.tail(2)
        distances_between_two_points = calculate_distance_between_two_points(last_two_rows.iloc[0], last_two_rows.iloc[1], G, filtered_df_within_circle)

        closest_points_indices = sorted(distances_between_two_points, key=distances_between_two_points.get)[:2]
        closest_points = filtered_df_within_circle.loc[closest_points_indices]
        min_perimeter_outlet_df = find_nearest_point(visited_points, closest_points)

        visited_points = pd.concat([visited_points, min_perimeter_outlet_df], ignore_index=True)
        filtered_sorted_df = group_1_df[~group_1_df['OutletID'].isin(visited_points['OutletID'])]
        filtered_sorted_df.info()
        cal_distance, cal_time, new_map = draw_optimal_path(visited_points, new_map, G, random_color, group_feature)
        total_distance += cal_distance
        total_time += cal_time
          
    return visited_points, new_map, total_distance, total_time

def Calculate_center_square(all_data_PG):
    # Gom nhóm dữ liệu theo cột "SRD" và tính tâm hình vuông cho mỗi nhóm
    grouped_tier_PG = all_data_PG.groupby('SRD')

    # Tính tâm hình vuông cho mỗi nhóm
    square_centers_PG = grouped_tier_PG.agg({
        'Latitude': 'mean',
        'Longitude': 'mean'
    })

    return square_centers_PG

def find_nearest_center_square(square_centers_PG, square_centers_B, all_data_B):
    # Lấy tọa độ của 4 tâm hình vuông đầu tiên từ square_centers_PG
    centers_PG = square_centers_PG.head(4)[['Latitude', 'Longitude']].values

    # Tạo một DataFrame mới chứa khoảng cách giữa các tâm trong square_centers_B và các tâm trong square_centers_PG
    distances = pd.DataFrame(index=square_centers_B.index, columns=['Distance'])

    for index, row in square_centers_B.iterrows():
        center_B = np.array([row['Latitude'], row['Longitude']])
        # Tính khoảng cách từ center_B đến các tâm trong centers_PG
        distances.loc[index, 'Distance'] = np.mean(np.sqrt(np.sum((centers_PG.astype(float) - center_B.astype(float))**2, axis=1)))
        
    # Tìm tâm trong square_centers_B có khoảng cách nhỏ nhất
    nearest_center_B = distances['Distance'].idxmin()
    
    # Tìm mọi thông tin Outlet mà có khoảng cách nhỏ nhất
    nearest_outlets_B_1 = all_data_B[all_data_B['SRD'] == nearest_center_B]
    
    return nearest_outlets_B_1

def find_Outlet_f4(square_centers_PG, all_data_PG):
    centers_PG = square_centers_PG.head(4)[['Latitude', 'Longitude']].values
    
    duplicates = []
    for index, row in square_centers_PG.iterrows():
        if [row['Latitude'], row['Longitude']] in centers_PG:
            duplicates.append((row['Latitude'], row['Longitude']))
            
    # Tạo DataFrame rỗng để chứa kết quả
    matched_rows = pd.DataFrame(columns=all_data_PG.columns)

    # Duyệt qua từng nhóm trong all_data_PG.groupby('SRD')
    for group_name, group_data in all_data_PG.groupby('SRD'):
        # Tính tâm hình vuông cho nhóm hiện tại
        group_center = (group_data['Latitude'].mean(), group_data['Longitude'].mean())
        
        # Kiểm tra xem tâm hình vuông của nhóm có trong danh sách duplicates không
        if group_center in duplicates:
            # Nếu có, thêm các dòng của nhóm này vào matched_rows
            matched_rows = pd.concat([matched_rows, group_data])

    return matched_rows   

def create_square_map2(group_1_df, min_lat, max_lat, min_lon, max_lon, baby_map):
    # Tính toán tọa độ của tâm hình vuông
    center_lat = (min_lat + max_lat) / 2
    center_lon = (min_lon + max_lon) / 2

    # Chọn màu ngẫu nhiên từ danh sách các màu
    colors = ['black', 'beige', 'lightblue', 'gray', 'blue', 'darkred', 'lightgreen', 'purple', 'red', 'green', 'lightred', 'darkblue', 'darkpurple', 'cadetblue', 'orange', 'pink', 'lightgray', 'darkgreen', 'pink', 'yellow', 'purple']

    random_color = random.choice(colors)
    
    # Thêm marker cho mỗi điểm trong group_1_df
    for index, row in group_1_df.iterrows():
        popup_content = f"Order: {index+1}<br>OutletID: {row['OutletID']}<br>OutletName: {row['OutletName']}<br>Latitude: {row['Latitude']}<br>Longitude: {row['Longitude']}<br>tier: {row['tier']}<br>sale: {row['sale']}<br>Flow: {row['Flow']}<br>List: {row['List']}<br>Sequence: {row['Sequence']}"
        folium.Marker(location=[row['Latitude'], row['Longitude']], popup=folium.Popup(popup_content, max_width=300), icon=folium.Icon(color=random_color)).add_to(baby_map)

    # Tạo hình vuông bao quanh các điểm với màu được chọn ngẫu nhiên
    folium.Rectangle(bounds=[(min_lat, min_lon), (max_lat, max_lon)], color=random_color, fill=True, fill_opacity=0.2).add_to(baby_map)

    return baby_map

def phanluong_for1sale(PG_data, S_data, B_data, i, map):
    week1_sale_1 = pd.DataFrame()
    week2_sale_1 = pd.DataFrame()
    week3_sale_1 = pd.DataFrame()
    week4_sale_1 = pd.DataFrame()

    if len(PG_data) > 0:
        square_centers_PG = Calculate_center_square(PG_data)
        if len(B_data) > 0:
            square_centers_B = Calculate_center_square(B_data)
            if len(S_data) > 0:
                square_centers_S = Calculate_center_square(S_data)
                nearest_outlets_B_1 = find_nearest_center_square(square_centers_PG, square_centers_B, B_data)
                nearest_outlets_S_1 = find_nearest_center_square(square_centers_PG, square_centers_S, S_data)
                matched_rows = find_Outlet_f4(square_centers_PG, PG_data)
                week1_sale_1 = pd.concat([matched_rows, nearest_outlets_B_1])
                week1_sale_1 = pd.concat([week1_sale_1, nearest_outlets_S_1])

                selected_outlet_ids = week1_sale_1['OutletID'].tolist()
                B_data = B_data[~B_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(B_data) > 0:
                    square_centers_B = Calculate_center_square(B_data)
                    nearest_outlets_B_2 = find_nearest_center_square(square_centers_PG, square_centers_B, B_data)
                    week2_sale_1 = pd.concat([matched_rows, nearest_outlets_S_1])
                    week2_sale_1 = pd.concat([week2_sale_1, nearest_outlets_B_2])
                else: 
                    week2_sale_1 = pd.concat([matched_rows, nearest_outlets_S_1])      
                
                selected_outlet_ids = week2_sale_1['OutletID'].tolist()
                S_data = S_data[~S_data['OutletID'].isin(selected_outlet_ids)]
                B_data = B_data[~B_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(B_data) > 0:
                    square_centers_B = Calculate_center_square(B_data)
                    if len(S_data) > 0:
                        square_centers_S = Calculate_center_square(S_data)
                        nearest_outlets_B_3 = find_nearest_center_square(square_centers_PG, square_centers_B, B_data)
                        nearest_outlets_S_2 = find_nearest_center_square(square_centers_PG, square_centers_S, S_data)
                        week3_sale_1 = pd.concat([matched_rows, nearest_outlets_B_3])
                        week3_sale_1 = pd.concat([week3_sale_1, nearest_outlets_S_2])
                    else:
                        nearest_outlets_B_3 = find_nearest_center_square(square_centers_PG, square_centers_B, B_data)
                        week3_sale_1 = pd.concat([matched_rows, nearest_outlets_B_3])
                else:
                    if len(S_data) > 0:
                        square_centers_S = Calculate_center_square(S_data)
                        nearest_outlets_S_2 = find_nearest_center_square(square_centers_PG, square_centers_S, S_data)
                        week3_sale_1 = pd.concat([matched_rows, nearest_outlets_S_2])
                    else:
                        week3_sale_1 = matched_rows.copy()
                
                selected_outlet_ids = week3_sale_1['OutletID'].tolist()
                B_data = B_data[~B_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(B_data) > 0:
                    square_centers_B = Calculate_center_square(B_data)
                    if len(S_data) > 0:
                        nearest_outlets_B_4 = find_nearest_center_square(square_centers_PG, square_centers_B, B_data)
                        week4_sale_1 = pd.concat([matched_rows, nearest_outlets_S_2])
                        week4_sale_1 = pd.concat([week4_sale_1, nearest_outlets_B_4])
                    else:
                        nearest_outlets_B_4 = find_nearest_center_square(square_centers_PG, square_centers_B, B_data)
                        week4_sale_1 = pd.concat([matched_rows, nearest_outlets_B_4])
                else:
                    if len(S_data) > 0:
                        week4_sale_1 = pd.concat([matched_rows, nearest_outlets_S_2])
                    else:
                        week4_sale_1 = matched_rows.copy()  
            else:
                nearest_outlets_B_1 = find_nearest_center_square(square_centers_PG, square_centers_B, B_data)
                matched_rows = find_Outlet_f4(square_centers_PG, PG_data)
                week1_sale_1 = pd.concat([matched_rows, nearest_outlets_B_1])

                selected_outlet_ids = week1_sale_1['OutletID'].tolist()
                B_data = B_data[~B_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(B_data) > 0:
                    square_centers_B = Calculate_center_square(B_data)
                    nearest_outlets_B_2 = find_nearest_center_square(square_centers_PG, square_centers_B, B_data)
                    week2_sale_1 = pd.concat([matched_rows, nearest_outlets_B_2])
                else:
                    week2_sale_1 = matched_rows.copy()
                
                selected_outlet_ids = week2_sale_1['OutletID'].tolist()
                B_data = B_data[~B_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(B_data) > 0:
                    square_centers_B = Calculate_center_square(B_data)
                    nearest_outlets_B_3 = find_nearest_center_square(square_centers_PG, square_centers_B, B_data)
                    week3_sale_1 = pd.concat([matched_rows, nearest_outlets_B_3])
                else:
                    week3_sale_1 = matched_rows.copy()
                    
                selected_outlet_ids = week3_sale_1['OutletID'].tolist()
                B_data = B_data[~B_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(B_data) > 0:
                    square_centers_B = Calculate_center_square(B_data)
                    nearest_outlets_B_4 = find_nearest_center_square(square_centers_PG, square_centers_B, B_data)
                    week4_sale_1 = pd.concat([matched_rows, nearest_outlets_B_4])
                else:
                    week4_sale_1 = matched_rows.copy()  
        else:
            if len(S_data) > 0:
                square_centers_S = Calculate_center_square(S_data)
                nearest_outlets_S_1 = find_nearest_center_square(square_centers_PG, square_centers_S, S_data)
                matched_rows = find_Outlet_f4(square_centers_PG, PG_data)
                week1_sale_1 = pd.concat([matched_rows, nearest_outlets_S_1])
                
                week2_sale_1 = week1_sale_1.copy()
                
                selected_outlet_ids = week2_sale_1['OutletID'].tolist()
                S_data = S_data[~S_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(S_data) > 0:
                    square_centers_S = Calculate_center_square(S_data)
                    nearest_outlets_S_2 = find_nearest_center_square(square_centers_PG, square_centers_S, S_data)
                    week3_sale_1 = pd.concat([matched_rows, nearest_outlets_S_2])
                else:
                    week3_sale_1 = matched_rows.copy()
                    
                week4_sale_1 = week3_sale_1.copy()                                 
    
            else:
                week1_sale_1 = PG_data.copy()
    
    if not week1_sale_1.empty:        
        week1_sale_1['sale'] = 'week1_6days_sale' + str(i)
        min_lat = week1_sale_1['Latitude'].min()
        max_lat = week1_sale_1['Latitude'].max()
        min_lon = week1_sale_1['Longitude'].min()
        max_lon = week1_sale_1['Longitude'].max()
        map = create_square_map2(week1_sale_1, min_lat, max_lat, min_lon, max_lon, map) 
    
    if not week2_sale_1.empty:
        week2_sale_1['sale'] = 'week2_6days_sale' + str(i)
        min_lat = week2_sale_1['Latitude'].min()
        max_lat = week2_sale_1['Latitude'].max()
        min_lon = week2_sale_1['Longitude'].min()
        max_lon = week2_sale_1['Longitude'].max()
        map = create_square_map2(week2_sale_1, min_lat, max_lat, min_lon, max_lon, map)

    if not week3_sale_1.empty:
        week3_sale_1['sale'] = 'week3_6days_sale' + str(i)
        min_lat = week3_sale_1['Latitude'].min()
        max_lat = week3_sale_1['Latitude'].max()
        min_lon = week3_sale_1['Longitude'].min()
        max_lon = week3_sale_1['Longitude'].max()
        map = create_square_map2(week3_sale_1, min_lat, max_lat, min_lon, max_lon, map)

    if not week4_sale_1.empty:
        week4_sale_1['sale'] = 'week4_6days_sale' + str(i)
        min_lat = week4_sale_1['Latitude'].min()
        max_lat = week4_sale_1['Latitude'].max()
        min_lon = week4_sale_1['Longitude'].min()
        max_lon = week4_sale_1['Longitude'].max()
        map = create_square_map2(week4_sale_1, min_lat, max_lat, min_lon, max_lon, map)
            
    return week1_sale_1, week2_sale_1, week3_sale_1, week4_sale_1, map

def run_for_range_n(PG_data, S_data, B_data, n, map):
    all_week_sales = []
    for i in range(1, n + 1):
        # Call phanluong_for1sale_5days for the current value of i
        week1_sale, week2_sale, week3_sale, week4_sale, map = phanluong_for1sale(PG_data, S_data, B_data, i, map)
        
        # Check if 'OutletID' column exists in any of the DataFrames
        outlet_ids = []
        for week_sale in [week1_sale, week2_sale, week3_sale, week4_sale]:
            if 'OutletID' in week_sale.columns:
                outlet_ids.extend(week_sale['OutletID'].tolist())
        
        # Update PG_data, S_data, B_data for the next iteration
        PG_data = PG_data[~PG_data['OutletID'].isin(outlet_ids)]
        B_data = B_data[~B_data['OutletID'].isin(outlet_ids)]
        S_data = S_data[~S_data['OutletID'].isin(outlet_ids)]
        
        # Thêm các bộ dữ liệu tuần vào danh sách all_week_sales_5days
        all_week_sales.append((week1_sale, week2_sale, week3_sale, week4_sale))

    return all_week_sales, map

def find_nearest_center_square_5days(square_centers_PG, square_centers_B, all_data_B):
    # Lấy tọa độ của 4 tâm hình vuông đầu tiên từ square_centers_PG
    centers_PG = square_centers_PG.head(3)[['Latitude', 'Longitude']].values

    # Tạo một DataFrame mới chứa khoảng cách giữa các tâm trong square_centers_B và các tâm trong square_centers_PG
    distances = pd.DataFrame(index=square_centers_B.index, columns=['Distance'])

    for index, row in square_centers_B.iterrows():
        center_B = np.array([row['Latitude'], row['Longitude']])
        # Tính khoảng cách từ center_B đến các tâm trong centers_PG
        distances.loc[index, 'Distance'] = np.mean(np.sqrt(np.sum((centers_PG.astype(float) - center_B.astype(float))**2, axis=1)))
        
    # Tìm tâm trong square_centers_B có khoảng cách nhỏ nhất
    nearest_center_B = distances['Distance'].idxmin()
    
    # Tìm mọi thông tin Outlet mà có khoảng cách nhỏ nhất
    nearest_outlets_B_1 = all_data_B[all_data_B['SRD'] == nearest_center_B]
    
    return nearest_outlets_B_1

def find_Outlet_f4_5days(square_centers_PG, all_data_PG):
    centers_PG = square_centers_PG.head(3)[['Latitude', 'Longitude']].values
    
    duplicates = []
    for index, row in square_centers_PG.iterrows():
        if [row['Latitude'], row['Longitude']] in centers_PG:
            duplicates.append((row['Latitude'], row['Longitude']))
            
    # Tạo DataFrame rỗng để chứa kết quả
    matched_rows = pd.DataFrame(columns=all_data_PG.columns)

    # Duyệt qua từng nhóm trong all_data_PG.groupby('SRD')
    for group_name, group_data in all_data_PG.groupby('SRD'):
        # Tính tâm hình vuông cho nhóm hiện tại
        group_center = (group_data['Latitude'].mean(), group_data['Longitude'].mean())
        
        # Kiểm tra xem tâm hình vuông của nhóm có trong danh sách duplicates không
        if group_center in duplicates:
            # Nếu có, thêm các dòng của nhóm này vào matched_rows
            matched_rows = pd.concat([matched_rows, group_data])

    return matched_rows 

def download_excel(dataframe, filename):
    # Tạo một đối tượng Workbook từ openpyxl
    wb = Workbook()
    ws = wb.active
    
    # Ghi tên các cột vào hàng đầu tiên
    for c_idx, col_name in enumerate(dataframe.columns, 1):
        ws.cell(row=1, column=c_idx, value=col_name)
    
    # Ghi dữ liệu từ DataFrame vào worksheet từ hàng thứ hai trở đi
    for r_idx, row in enumerate(dataframe.iterrows(), 2):  # Bắt đầu từ hàng thứ hai
        for c_idx, value in enumerate(row[1], 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Lưu workbook vào đối tượng BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    
    # Lấy nội dung từ đối tượng BytesIO và mã hóa nó thành base64
    excel_binary = excel_buffer.getvalue()
    b64 = base64.b64encode(excel_binary).decode()
    
    # Tạo liên kết để tải xuống file Excel với UTF-8 encoding
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;charset=utf-8;base64,{b64}" download="{filename}.xlsx">Download Excel</a>'
    
    return href

def get_html_from_map(new_map):
    """Get HTML string from folium.Map object"""
    tmpfile = BytesIO()
    new_map.save(tmpfile, close_file=False)
    html = tmpfile.getvalue().decode()
    return html

def phanluong_for1sale_5days(PG_data, S_data, B_data, i, map):
    week1_sale_1 = pd.DataFrame()
    week2_sale_1 = pd.DataFrame()
    week3_sale_1 = pd.DataFrame()
    week4_sale_1 = pd.DataFrame()

    if len(PG_data) > 0:
        square_centers_PG = Calculate_center_square(PG_data)
        if len(B_data) > 0:
            square_centers_B = Calculate_center_square(B_data)
            if len(S_data) > 0:
                square_centers_S = Calculate_center_square(S_data)
                nearest_outlets_B_1 = find_nearest_center_square_5days(square_centers_PG, square_centers_B, B_data)
                nearest_outlets_S_1 = find_nearest_center_square_5days(square_centers_PG, square_centers_S, S_data)
                matched_rows = find_Outlet_f4_5days(square_centers_PG, PG_data)
                week1_sale_1 = pd.concat([matched_rows, nearest_outlets_B_1])
                week1_sale_1 = pd.concat([week1_sale_1, nearest_outlets_S_1])

                selected_outlet_ids = week1_sale_1['OutletID'].tolist()
                B_data = B_data[~B_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(B_data) > 0:
                    square_centers_B = Calculate_center_square(B_data)
                    nearest_outlets_B_2 = find_nearest_center_square_5days(square_centers_PG, square_centers_B, B_data)
                    week2_sale_1 = pd.concat([matched_rows, nearest_outlets_S_1])
                    week2_sale_1 = pd.concat([week2_sale_1, nearest_outlets_B_2])
                else: 
                    week2_sale_1 = pd.concat([matched_rows, nearest_outlets_S_1])      
                
                selected_outlet_ids = week2_sale_1['OutletID'].tolist()
                S_data = S_data[~S_data['OutletID'].isin(selected_outlet_ids)]
                B_data = B_data[~B_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(B_data) > 0:
                    square_centers_B = Calculate_center_square(B_data)
                    if len(S_data) > 0:
                        square_centers_S = Calculate_center_square(S_data)
                        nearest_outlets_B_3 = find_nearest_center_square_5days(square_centers_PG, square_centers_B, B_data)
                        nearest_outlets_S_2 = find_nearest_center_square_5days(square_centers_PG, square_centers_S, S_data)
                        week3_sale_1 = pd.concat([matched_rows, nearest_outlets_B_3])
                        week3_sale_1 = pd.concat([week3_sale_1, nearest_outlets_S_2])
                    else:
                        nearest_outlets_B_3 = find_nearest_center_square_5days(square_centers_PG, square_centers_B, B_data)
                        week3_sale_1 = pd.concat([matched_rows, nearest_outlets_B_3])
                else:
                    if len(S_data) > 0:
                        square_centers_S = Calculate_center_square(S_data)
                        nearest_outlets_S_2 = find_nearest_center_square_5days(square_centers_PG, square_centers_S, S_data)
                        week3_sale_1 = pd.concat([matched_rows, nearest_outlets_S_2])
                    else:
                        week3_sale_1 = matched_rows.copy()
                
                selected_outlet_ids = week3_sale_1['OutletID'].tolist()
                B_data = B_data[~B_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(B_data) > 0:
                    square_centers_B = Calculate_center_square(B_data)
                    if len(S_data) > 0:
                        nearest_outlets_B_4 = find_nearest_center_square_5days(square_centers_PG, square_centers_B, B_data)
                        week4_sale_1 = pd.concat([matched_rows, nearest_outlets_S_2])
                        week4_sale_1 = pd.concat([week4_sale_1, nearest_outlets_B_4])
                    else:
                        nearest_outlets_B_4 = find_nearest_center_square_5days(square_centers_PG, square_centers_B, B_data)
                        week4_sale_1 = pd.concat([matched_rows, nearest_outlets_B_4])
                else:
                    if len(S_data) > 0:
                        week4_sale_1 = pd.concat([matched_rows, nearest_outlets_S_2])
                    else:
                        week4_sale_1 = matched_rows.copy()  
            else:
                nearest_outlets_B_1 = find_nearest_center_square_5days(square_centers_PG, square_centers_B, B_data)
                matched_rows = find_Outlet_f4_5days(square_centers_PG, PG_data)
                week1_sale_1 = pd.concat([matched_rows, nearest_outlets_B_1])

                selected_outlet_ids = week1_sale_1['OutletID'].tolist()
                B_data = B_data[~B_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(B_data) > 0:
                    square_centers_B = Calculate_center_square(B_data)
                    nearest_outlets_B_2 = find_nearest_center_square_5days(square_centers_PG, square_centers_B, B_data)
                    week2_sale_1 = pd.concat([matched_rows, nearest_outlets_B_2])
                else:
                    week2_sale_1 = matched_rows.copy()
                
                selected_outlet_ids = week2_sale_1['OutletID'].tolist()
                B_data = B_data[~B_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(B_data) > 0:
                    square_centers_B = Calculate_center_square(B_data)
                    nearest_outlets_B_3 = find_nearest_center_square_5days(square_centers_PG, square_centers_B, B_data)
                    week3_sale_1 = pd.concat([matched_rows, nearest_outlets_B_3])
                else:
                    week3_sale_1 = matched_rows.copy()
                    
                selected_outlet_ids = week3_sale_1['OutletID'].tolist()
                B_data = B_data[~B_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(B_data) > 0:
                    square_centers_B = Calculate_center_square(B_data)
                    nearest_outlets_B_4 = find_nearest_center_square_5days(square_centers_PG, square_centers_B, B_data)
                    week4_sale_1 = pd.concat([matched_rows, nearest_outlets_B_4])
                else:
                    week4_sale_1 = matched_rows.copy()  
        else:
            if len(S_data) > 0:
                square_centers_S = Calculate_center_square(S_data)
                nearest_outlets_S_1 = find_nearest_center_square_5days(square_centers_PG, square_centers_S, S_data)
                matched_rows = find_Outlet_f4_5days(square_centers_PG, PG_data)
                week1_sale_1 = pd.concat([matched_rows, nearest_outlets_S_1])
                
                week2_sale_1 = week1_sale_1.copy()
                
                selected_outlet_ids = week2_sale_1['OutletID'].tolist()
                S_data = S_data[~S_data['OutletID'].isin(selected_outlet_ids)]
                
                if len(S_data) > 0:
                    square_centers_S = Calculate_center_square(S_data)
                    nearest_outlets_S_2 = find_nearest_center_square_5days(square_centers_PG, square_centers_S, S_data)
                    week3_sale_1 = pd.concat([matched_rows, nearest_outlets_S_2])
                else:
                    week3_sale_1 = matched_rows.copy()
                    
                week4_sale_1 = week3_sale_1.copy()
            else:
                week1_sale_1 = PG_data.copy()
    
    if not week1_sale_1.empty:        
        week1_sale_1['sale'] = 'week1_5days_sale' + str(i)
        min_lat = week1_sale_1['Latitude'].min()
        max_lat = week1_sale_1['Latitude'].max()
        min_lon = week1_sale_1['Longitude'].min()
        max_lon = week1_sale_1['Longitude'].max()
        map = create_square_map2(week1_sale_1, min_lat, max_lat, min_lon, max_lon, map) 
    
    if not week2_sale_1.empty:
        week2_sale_1['sale'] = 'week2_5days_sale' + str(i)
        min_lat = week2_sale_1['Latitude'].min()
        max_lat = week2_sale_1['Latitude'].max()
        min_lon = week2_sale_1['Longitude'].min()
        max_lon = week2_sale_1['Longitude'].max()
        map = create_square_map2(week2_sale_1, min_lat, max_lat, min_lon, max_lon, map)

    if not week3_sale_1.empty:
        week3_sale_1['sale'] = 'week3_5days_sale' + str(i)
        min_lat = week3_sale_1['Latitude'].min()
        max_lat = week3_sale_1['Latitude'].max()
        min_lon = week3_sale_1['Longitude'].min()
        max_lon = week3_sale_1['Longitude'].max()
        map = create_square_map2(week3_sale_1, min_lat, max_lat, min_lon, max_lon, map)

    if not week4_sale_1.empty:
        week4_sale_1['sale'] = 'week4_5days_sale' + str(i)
        min_lat = week4_sale_1['Latitude'].min()
        max_lat = week4_sale_1['Latitude'].max()
        min_lon = week4_sale_1['Longitude'].min()
        max_lon = week4_sale_1['Longitude'].max()
        map = create_square_map2(week4_sale_1, min_lat, max_lat, min_lon, max_lon, map)
            
    return week1_sale_1, week2_sale_1, week3_sale_1, week4_sale_1, map

def run_for_range_n_5days(PG_data, S_data, B_data, n, map):
    all_week_sales_5days = []
    for i in range(1, n + 1):
        # Call phanluong_for1sale_5days for the current value of i
        week1_sale, week2_sale, week3_sale, week4_sale, map = phanluong_for1sale_5days(PG_data, S_data, B_data, i, map)
        
        # Check if 'OutletID' column exists in any of the DataFrames
        outlet_ids = []
        for week_sale in [week1_sale, week2_sale, week3_sale, week4_sale]:
            if 'OutletID' in week_sale.columns:
                outlet_ids.extend(week_sale['OutletID'].tolist())
        
        # Update PG_data, S_data, B_data for the next iteration
        PG_data = PG_data[~PG_data['OutletID'].isin(outlet_ids)]
        B_data = B_data[~B_data['OutletID'].isin(outlet_ids)]
        S_data = S_data[~S_data['OutletID'].isin(outlet_ids)]
        
        # Thêm các bộ dữ liệu tuần vào danh sách all_week_sales_5days
        all_week_sales_5days.append((week1_sale, week2_sale, week3_sale, week4_sale))

    return all_week_sales_5days, map

def calculate_service_time(row):
    if row['tier'] == 'P':
        return 30
    elif row['tier'] == 'G':
        return 20
    elif row['tier'] == 'S':
        return 15
    elif row['tier'] == 'B':
        return 10
    
def main():
    st.markdown("<h1 style='text-align: center; font-size: 55px;'>Traveling Salesman Problem</h1>", unsafe_allow_html=True)

    # Upload files
    st.header("1. Upload Excel File")

    # Kiểm tra số lượng file đã tải lên
    uploaded_files = st.file_uploader("Upload Excel file", type=["xlsx"], accept_multiple_files=True)
    
    dataframes = {}
    data = None
    final_df1 = None

    if uploaded_files:
        dataframes, data = process_uploaded_files(uploaded_files)

        no_outlet_f4 = st.slider("Select number outlet F4:", 0, 40, 24, 1)
        st.text(f"Selected number for F4: {no_outlet_f4}")

        no_outlet_f2 = st.slider("Select number outlet F2:", 0, 40, 30, 1)
        st.text(f"Selected number for F2: {no_outlet_f2}")

        no_outlet_f1 = st.slider("Select number outlet F1:", 0, 40, 30, 1)
        st.text(f"Selected number for F1: {no_outlet_f1}")

        no_sale6days = st.slider("Select number sale 6 days:", 0, 40, 2, 1)
        st.text(f"Selected number for sale 6 days: {no_sale6days}")
        
        no_sale5days = st.slider("Select number sale 5 days:", 0, 40, 2, 1)
        st.text(f"Selected number for sale 5 days: {no_sale5days}")        
                        
        # Tạo text input cho vị trí (location)
        location = st.text_input("Nhập vị trí (location):")

        # network_type
        network_type = 'bike'
        
        if location:
            st.header("2. Result")
            
            G = ox.graph_from_place(location, network_type=network_type)
            st.text("Loaded Map Done")
            
            data['Longitude'] = data['Longitude'].astype(float)
            data['Latitude'] = data['Latitude'].astype(float)
            
            cleaned_latitude = remove_outliers_iqr(data['Latitude'])
            cleaned_longitude = remove_outliers_iqr(data['Longitude'])

            # Làm sạch dữ liệu
            cleaned_data = data[(data['Latitude'].isin(cleaned_latitude)) & (data['Longitude'].isin(cleaned_longitude))]

            data_PG = cleaned_data[(cleaned_data['tier'] == 'P') | (cleaned_data['tier'] == 'G')]
            data_S = cleaned_data[cleaned_data['tier'] == 'S']
            data_B = cleaned_data[cleaned_data['tier'] == 'B']

            data_PG['flow'] = 'F4'
            data_S['flow'] = 'F2'
            data_B['flow'] = 'F1'
            
            # Create a new map for optimized paths
            map = folium.Map(location=[cleaned_data.iloc[0]['Latitude'], cleaned_data.iloc[0]['Longitude']], zoom_start=12)

            # Iterate through each point in group_1_df
            for index, row in cleaned_data.iterrows():
                # Add a marker for each point with a popup displaying its order and information
                popup_content = f"Index: {index+1}<br>OutletID: {row['OutletID']}<br>OutletName: {row['OutletName']}<br>Latitude: {row['Latitude']}<br>Longitude: {row['Longitude']}<br>tier: {row['tier']}"
                folium.Marker(location=[row['Latitude'], row['Longitude']], popup=folium.Popup(popup_content, max_width=300)).add_to(map)
                
            random_color = 'gray'
            all_data_PG, map = Create_square(data_PG, no_outlet_f4, map, random_color)
            all_data_PG = Create_RD(all_data_PG)  
            
            random_color = 'green'
            all_data_S, map = Create_square(data_S, no_outlet_f2, map, random_color)
            all_data_S = Create_RD(all_data_S)
            
            random_color = 'pink'
            all_data_B, map = Create_square(data_B, no_outlet_f1, map, random_color)
            all_data_B = Create_RD(all_data_B)

            all_data_PG = all_data_PG.drop(columns=['distance_to_min', 'distance_to_closest'])
            all_data_S = all_data_S.drop(columns=['distance_to_min', 'distance_to_closest'])
            all_data_B = all_data_B.drop(columns=['distance_to_min', 'distance_to_closest'])  
            
            all_data_PG['flow'] = 'F4'
            all_data_PG['Flow'] = all_data_PG['flow'] + '-' + all_data_PG['SRD'].astype(str)
            all_data_S['flow'] = 'F2'
            all_data_S['Flow'] = all_data_S['flow'] + '-' + all_data_S['SRD'].astype(str)
            all_data_B['flow'] = 'F1'
            all_data_B['Flow'] = all_data_B['flow'] + '-' + all_data_B['SRD'].astype(str)


            gom_data = pd.DataFrame(columns=all_data_B.columns)
            gom_data = pd.concat([gom_data, all_data_PG], ignore_index=True)
            gom_data = pd.concat([gom_data, all_data_S], ignore_index=True)
            gom_data = pd.concat([gom_data, all_data_B], ignore_index=True)

            # Create a new map for optimized paths
            test_map = folium.Map(location=[gom_data.iloc[0]['Latitude'], gom_data.iloc[0]['Longitude']], zoom_start=12)

            # Iterate through each point in group_1_df
            for index, row in gom_data.iterrows():
                # Add a marker for each point with a popup displaying its order and information
                popup_content = f"Index: {index+1}<br>OutletID: {row['OutletID']}<br>OutletName: {row['OutletName']}<br>Latitude: {row['Latitude']}<br>Longitude: {row['Longitude']}<br>flow: {row['flow']}<br>Flow: {row['Flow']}"
                folium.Marker(location=[row['Latitude'], row['Longitude']], popup=folium.Popup(popup_content, max_width=300)).add_to(test_map) 
                
            # Find numer of loop for each tier
            sovongchay_PG = all_data_PG['SRD'].value_counts().index[-1] + 1
            sovongchay_S = all_data_S['SRD'].value_counts().index[-1] + 1
            sovongchay_B = all_data_B['SRD'].value_counts().index[-1] + 1

            # Initialize lists for PG and S groups
            visited_points_listf2 = []
            total_distance_listf2 = []
            total_time_listf2 = []
            visited_points_listf4 = []
            total_distance_listf4 = []
            total_time_listf4 = []
            visited_points_listf1 = []
            total_distance_listf1 = []
            total_time_listf1 = []

            # Process data for PG groups
            for i in range(1, sovongchay_PG):
                print('Đang xử lý lần thứ ' + str(i))
                # Filter data for the current group (i)
                group_df = all_data_PG[all_data_PG['SRD'] == i]

                colors = ['black', 'lightblue', 'gray', 'blue', 'lightgreen', 'purple', 'red', 'green', 'white', 'darkblue', 'orange', 'pink', 'yellow']
                random_color = random.choice(colors)
                
                # Call create_path and store returned values
                visited_points_i, test_map, total_distance, total_time = create_path(group_df, G, test_map, random_color, i)

                # Append values to PG lists
                visited_points_listf4.append(visited_points_i)
                total_distance_listf4.append(total_distance)
                total_time_listf4.append(total_time)

                print('Hoàn thành xử lý lần thứ ' + str(i))

            # Process data for S groups
            for i in range(1, sovongchay_S):
                print('Đang xử lý lần thứ ' + str(i))
                # Filter data for the current group (i)
                group_df2 = all_data_S[all_data_S['SRD'] == i]

                colors = ['black', 'lightblue', 'gray', 'blue', 'lightgreen', 'purple', 'red', 'green', 'white', 'darkblue', 'orange', 'pink', 'yellow']
                random_color = random.choice(colors)
                
                # Call create_path and store returned values
                visited_points_2, test_map, total_distance2, total_time2 = create_path(group_df2, G, test_map, random_color, i)

                # Append values to S lists
                visited_points_listf2.append(visited_points_2)
                total_distance_listf2.append(total_distance2)
                total_time_listf2.append(total_time2)

                print('Hoàn thành xử lý lần thứ ' + str(i))
                
            # Process data for B groups
            for i in range(1, sovongchay_B):
                print('Đang xử lý lần thứ ' + str(i))
                # Filter data for the current group (i)
                group_df3 = all_data_B[all_data_B['SRD'] == i]

                colors = ['black', 'lightblue', 'gray', 'blue', 'lightgreen', 'purple', 'red', 'green', 'white', 'darkblue', 'orange', 'pink', 'yellow']
                random_color = random.choice(colors)
                
                # Call create_path and store returned values
                visited_points_3, test_map, total_distance3, total_time3 = create_path(group_df3, G, test_map, random_color, i)

                # Append values to S lists
                visited_points_listf1.append(visited_points_3)
                total_distance_listf1.append(total_distance3)
                total_time_listf1.append(total_time3)

                print('Hoàn thành xử lý lần thứ ' + str(i))    

            # Create a Layer Control for PG groups
            layer_control = folium.LayerControl().add_to(test_map)

            # Khởi tạo DataFrame rỗng
            danhsach_f4 = pd.DataFrame()

            # Duyệt qua từng DataFrame trong visited_points_list
            for i, df in enumerate(visited_points_listf4):
                # Tạo cột 'List' và gán giá trị là số thứ tự của df + 1
                df['List'] = i + 1
                # Tạo cột 'Sequence' và gán giá trị từ 1 đến chiều dài của df
                df['Sequence'] = range(1, len(df) + 1)
                # Kết hợp DataFrame hiện tại vào danhsach_f4
                danhsach_f4 = pd.concat([danhsach_f4, df], ignore_index=True)

            # Khởi tạo DataFrame rỗng
            danhsach_f2 = pd.DataFrame()

            # Duyệt qua từng DataFrame trong visited_points_list
            for i, df in enumerate(visited_points_listf2):
                # Tạo cột 'List' và gán giá trị là số thứ tự của df + 1
                df['List'] = i + 1
                # Tạo cột 'Sequence' và gán giá trị từ 1 đến chiều dài của df
                df['Sequence'] = range(1, len(df) + 1)
                # Kết hợp DataFrame hiện tại vào danhsach_f2
                danhsach_f2 = pd.concat([danhsach_f2, df], ignore_index=True) 
                
            # Khởi tạo DataFrame rỗng
            danhsach_f1 = pd.DataFrame()

            # Duyệt qua từng DataFrame trong visited_points_list
            for i, df in enumerate(visited_points_listf1):
                # Tạo cột 'List' và gán giá trị là số thứ tự của df + 1
                df['List'] = i + 1
                # Tạo cột 'Sequence' và gán giá trị từ 1 đến chiều dài của df
                df['Sequence'] = range(1, len(df) + 1)
                # Kết hợp DataFrame hiện tại vào danhsach_f1
                danhsach_f1 = pd.concat([danhsach_f1, df], ignore_index=True)
                
            danhsach_f4 = danhsach_f4.drop(columns=['WardName', 'DistrictName','ProvinceName', 'flow', 'distance_to_point'])
            danhsach_f2 = danhsach_f2.drop(columns=['WardName', 'DistrictName','ProvinceName', 'flow', 'distance_to_point'])
            danhsach_f1 = danhsach_f1.drop(columns=['WardName', 'DistrictName','ProvinceName', 'flow', 'distance_to_point'])
            
            result_dff4 = pd.DataFrame()
            result_dff2 = pd.DataFrame()
            result_dff1 = pd.DataFrame()
            
            # Tạo DataFrame rỗng
            result_dff4 = pd.DataFrame(columns=['Flow', 'Total Distance (km)', 'Travel_Total_Hours', 'Travel_Total_Minutes', 'Travel_Total_Seconds'])

            # Duyệt qua từng phần tử trong total_distance_list và total_time_list
            for i in range(len(total_distance_listf4)):
                total_distance_km = total_distance_listf4[i] / 1000
                total_time = total_time_listf4[i]
                total_hours = total_time // 3600
                remaining_seconds = total_time % 3600
                total_minutes = remaining_seconds // 60
                total_seconds = remaining_seconds % 60
                
                # Thêm dòng vào DataFrame
                result_dff4 = pd.concat([result_dff4, pd.DataFrame({
                    'Flow': ["F4-" + str(i+1)],
                    'Total Distance (km)': [total_distance_km],
                    'Travel_Total_Hours': [int(total_hours)],
                    'Travel_Total_Minutes': [int(total_minutes)],
                    'Travel_Total_Seconds': [int(total_seconds)]
                })], ignore_index=True)                                        

            # Tạo DataFrame rỗng
            result_dff2 = pd.DataFrame(columns=['Flow', 'Total Distance (km)', 'Travel_Total_Hours', 'Travel_Total_Minutes', 'Travel_Total_Seconds'])

            # Duyệt qua từng phần tử trong total_distance_list và total_time_list
            for i in range(len(total_distance_listf2)):
                total_distance_km = total_distance_listf2[i] / 1000
                total_time = total_time_listf2[i]
                total_hours = total_time // 3600
                remaining_seconds = total_time % 3600
                total_minutes = remaining_seconds // 60
                total_seconds = remaining_seconds % 60
                
                # Thêm dòng vào DataFrame
                result_dff2 = pd.concat([result_dff2, pd.DataFrame({
                    'Flow': ["F2-" + str(i+1)],
                    'Total Distance (km)': [total_distance_km],
                    'Travel_Total_Hours': [int(total_hours)],
                    'Travel_Total_Minutes': [int(total_minutes)],
                    'Travel_Total_Seconds': [int(total_seconds)]
                })], ignore_index=True)

            # Tạo DataFrame rỗng
            result_dff1 = pd.DataFrame(columns=['Flow', 'Total Distance (km)', 'Travel_Total_Hours', 'Travel_Total_Minutes', 'Travel_Total_Seconds'])

            # Duyệt qua từng phần tử trong total_distance_list và total_time_list
            for i in range(len(total_distance_listf1)):
                total_distance_km = total_distance_listf1[i] / 1000
                total_time = total_time_listf1[i]
                total_hours = total_time // 3600
                remaining_seconds = total_time % 3600
                total_minutes = remaining_seconds // 60
                total_seconds = remaining_seconds % 60
                
                # Thêm dòng vào DataFrame
                result_dff1 = pd.concat([result_dff1, pd.DataFrame({
                    'Flow': ["F1-" + str(i+1)],
                    'Total Distance (km)': [total_distance_km],
                    'Travel_Total_Hours': [int(total_hours)],
                    'Travel_Total_Minutes': [int(total_minutes)],
                    'Travel_Total_Seconds': [int(total_seconds)]
                })], ignore_index=True)

            PG_data = danhsach_f4.copy()
            S_data = danhsach_f2.copy()
            B_data = danhsach_f1.copy()

            all_week_sales, test_map = run_for_range_n(PG_data, S_data, B_data, no_sale6days, test_map)

            # Initialize an empty list to store all week sales dataframes
            all_weeks = []

            # Iterate through all_week_sales and concatenate the DataFrames for each week
            for week_sale in all_week_sales:
                all_weeks.extend(list(week_sale))

            # Concatenate all DataFrames in all_weeks list
            all_weeks_df = pd.concat(all_weeks, ignore_index=True)

            selected_outlet_ids = all_weeks_df['OutletID'].tolist()
            PG_data = PG_data[~PG_data['OutletID'].isin(selected_outlet_ids)]
            B_data = B_data[~B_data['OutletID'].isin(selected_outlet_ids)]
            S_data = S_data[~S_data['OutletID'].isin(selected_outlet_ids)]
            
            all_week_sales_5days, test_map = run_for_range_n_5days(PG_data, S_data, B_data, no_sale5days, test_map)
            
            # Initialize an empty list to store all week sales dataframes
            all_weeks_5days = []

            # Iterate through all_week_sales and concatenate the DataFrames for each week
            for week_sale in all_week_sales_5days:
                all_weeks_5days.extend(list(week_sale))

            # Concatenate all DataFrames in all_weeks_5days list
            all_weeks_df_5days = pd.concat(all_weeks_5days, ignore_index=True)
            
            all_weeks_df_test = all_weeks_df.copy()
            all_weeks_df_5days_test = all_weeks_df_5days.copy()

            df = pd.DataFrame(all_weeks_df_test)

            flow_counts = df.groupby('sale')['Flow'].value_counts()

            df2 = pd.DataFrame(all_weeks_df_5days_test)

            flow_counts2 = df2.groupby('sale')['Flow'].value_counts()
            
            all_weeks_df_test['service_time'] = all_weeks_df_test.apply(calculate_service_time, axis=1)
            all_weeks_df_5days_test['service_time'] = all_weeks_df_5days_test.apply(calculate_service_time, axis=1)
            
            # Nhóm theo cột "sale" và "Flow" và tính tổng service_time
            total_service_time_by_sale_flow = all_weeks_df_test.groupby(['sale', 'Flow'])['service_time'].sum().reset_index()

            # Chuyển đổi tổng service time từ phút sang giờ và phút
            total_service_time_by_sale_flow['Service_hours'] = total_service_time_by_sale_flow['service_time'] // 60
            total_service_time_by_sale_flow['Service_minutes'] = total_service_time_by_sale_flow['service_time'] % 60

            # Xuất ra DataFrame với 4 cột
            result_df = total_service_time_by_sale_flow[['sale', 'Flow', 'Service_hours', 'Service_minutes']]
            
            # Lấy các giá trị duy nhất từ các cột 'Flow', 'hours', 'minutes'
            distinct_values = result_df[['Flow', 'Service_hours', 'Service_minutes']].drop_duplicates()

            total_service_time_by_sale_flow = all_weeks_df_5days_test.groupby(['sale', 'Flow'])['service_time'].sum().reset_index()

            # Chuyển đổi tổng service time từ phút sang giờ và phút
            total_service_time_by_sale_flow['Service_hours'] = total_service_time_by_sale_flow['service_time'] // 60
            total_service_time_by_sale_flow['Service_minutes'] = total_service_time_by_sale_flow['service_time'] % 60

            # Xuất ra DataFrame với 4 cột
            result_df = total_service_time_by_sale_flow[['sale', 'Flow', 'Service_hours', 'Service_minutes']]            

            distinct_values2 = result_df[['Flow', 'Service_hours', 'Service_minutes']].drop_duplicates()

            merged_df_6days_f4 = pd.DataFrame()
            merged_df_6days_f2 = pd.DataFrame()
            merged_df_6days_f1 = pd.DataFrame()   
            
            merged_df_6days_f4 = pd.merge(result_dff4, distinct_values, on='Flow', how='right')
            merged_df_6days_f4 = merged_df_6days_f4.dropna()
            merged_df_6days_f2 = pd.merge(result_dff2, distinct_values, on='Flow', how='right')
            merged_df_6days_f2 = merged_df_6days_f2.dropna()
            merged_df_6days_f1 = pd.merge(result_dff1, distinct_values, on='Flow', how='right')
            merged_df_6days_f1 = merged_df_6days_f1.dropna()

            merged_df_5days_f4 = pd.DataFrame()
            merged_df_5days_f2 = pd.DataFrame()
            merged_df_5days_f1 = pd.DataFrame()    
            
            merged_df_5days_f4 = pd.merge(result_dff4, distinct_values2, on='Flow', how='right')
            merged_df_5days_f4 = merged_df_5days_f4.dropna()
            merged_df_5days_f2 = pd.merge(result_dff2, distinct_values2, on='Flow', how='right')
            merged_df_5days_f2 = merged_df_5days_f2.dropna()
            merged_df_5days_f1 = pd.merge(result_dff1, distinct_values2, on='Flow', how='right')
            merged_df_5days_f1 = merged_df_5days_f1.dropna()

            sales = flow_counts.index.get_level_values(0)
            flows = flow_counts.index.get_level_values(1)

            # Tạo DataFrame mới từ dữ liệu đã trích xuất và cột 'count'
            df3 = pd.DataFrame({'sale': sales, 'Flow': flows, 'count': flow_counts.values})

            sales2 = flow_counts2.index.get_level_values(0)
            flows2 = flow_counts2.index.get_level_values(1)

            # Tạo DataFrame mới từ dữ liệu đã trích xuất và cột 'count'
            df4 = pd.DataFrame({'sale': sales2, 'Flow': flows2, 'count': flow_counts2.values})
            
            st.header("Danh sách Week cho salesman 6 days")
            href_csv = download_excel(all_weeks_df_test, "all_weeks_df_test")
            st.markdown(href_csv, unsafe_allow_html=True)                                                                                                                                                                                            

            st.header("Danh sách Week cho salesman 5 days")
            href2_csv = download_excel(all_weeks_df_5days_test, "all_weeks_df_5days_test")
            st.markdown(href2_csv, unsafe_allow_html=True)    

            st.header("Thời gian f4 cho salesman 6 days")
            href2_csv = download_excel(merged_df_6days_f4, "merged_df_6days_f4")
            st.markdown(href2_csv, unsafe_allow_html=True)
                  
            st.header("Thời gian f4 cho salesman 5 days")
            href2_csv = download_excel(merged_df_5days_f4, "merged_df_5days_f4")
            st.markdown(href2_csv, unsafe_allow_html=True)
            
            st.header("Thời gian f2 cho salesman 6 days")
            href2_csv = download_excel(merged_df_6days_f2, "merged_df_6days_f2")
            st.markdown(href2_csv, unsafe_allow_html=True)  

            st.header("Thời gian f2 cho salesman 5 days")
            href2_csv = download_excel(merged_df_5days_f2, "merged_df_5days_f2")
            st.markdown(href2_csv, unsafe_allow_html=True)                

            st.header("Thời gian f1 cho salesman 6 days")
            href2_csv = download_excel(merged_df_6days_f1, "merged_df_6days_f1")
            st.markdown(href2_csv, unsafe_allow_html=True) 
            
            st.header("Thời gian f1 cho salesman 5 days")  
            href2_csv = download_excel(merged_df_5days_f1, "merged_df_5days_f1")
            st.markdown(href2_csv, unsafe_allow_html=True) 
  
            st.header("Thống kê số lượng Outlet cho salesman 6 days")
            href2_csv = download_excel(df3, "df3")
            st.markdown(href2_csv, unsafe_allow_html=True)   
            
            st.header("Thống kê số lượng Outlet cho salesman 5 days") 
            href2_csv = download_excel(df4, "df4")
            st.markdown(href2_csv, unsafe_allow_html=True)  
            
            st.header("Map (đường đi + gom nhóm theo hình vuông)") 
            html = get_html_from_map(test_map)
            b64 = base64.b64encode(html.encode()).decode()
            href = f'<a href="data:text/html;base64,{b64}" download="map.html">Download Map</a>'
            st.markdown(href, unsafe_allow_html=True)  
            # folium_static(test_map)
                        
if __name__ == '__main__':
    main()   